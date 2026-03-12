#!/usr/bin/env python3
"""
SSP-SP Delegacia-level roubo_outros scraper — SP Capital
=========================================================

The old ASP.NET Pesquisa.aspx page is gone (HTTP 404 since ~2023).
The new SSP portal serves bulk incident-level XLSX files at:
  https://www.ssp.sp.gov.br/assets/estatistica/transparencia/spDados/SPDadosCriminais_YYYY.xlsx

Each file contains individual Boletim de Ocorrência records with columns:
  NOME_DEPARTAMENTO          – 'DECAP' for São Paulo capital
  NOME_DELEGACIA             – precinct that registered the BO
  NOME_DELEGACIA_CIRCUNSCRIÇÃO – precinct with territorial jurisdiction (preferred for geo analysis)
  NATUREZA_APURADA           – crime category (e.g. 'ROUBO - OUTROS')
  ANO_ESTATISTICA / MES_ESTATISTICA

Files available (as of March 2026):
  2022 – JAN-JUN and JUL-DEZ sheets (~176 MB)
  2023 – JAN-JUN and JUL-DEZ sheets (~197 MB)
  2024 – JAN-JUN and JUL-DEZ sheets (~177 MB)
  2025 – (partial year)
  2026 – (partial year)

Usage
-----
  python3 scrape_ssp_delegacia.py

Output
------
  data/roubo_outros_por_delegacia.csv   – annual aggregated counts
  data/roubo_outros_mensal.csv          – monthly granularity

Delegacia → IBGE district mapping
----------------------------------
The delegacias are named like "03º D.P. CAMPOS ELISEOS".  A CSV of delegacia
names is written to data/sp_capital_delegacias.csv so you can manually or
fuzzy-match them to the 96 IBGE subprefectura districts.

NOTE: The XLSX files are large (~180-200 MB each). Set DOWNLOAD_FRESH=False
      to re-use cached files if you have already downloaded them.
"""

import os
import io
import csv
import json
import time
import logging
import urllib.request
from collections import defaultdict

import openpyxl

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BASE_URL = "https://www.ssp.sp.gov.br"
MANIFEST_URL = (
    f"{BASE_URL}/assets/estatistica/transparencia/spDados160_516.json"
)

OUTPUT_DIR = os.path.join(os.path.dirname(__file__))
CACHE_DIR = "/tmp/ssp_xlsx_cache"

# Set to False to skip re-downloading already cached files
DOWNLOAD_FRESH = True

# Which years to process
TARGET_YEARS = [2021, 2022, 2023, 2024]

# Department code for São Paulo capital
DECAP = "DECAP"

# The crime category we want (NATUREZA_APURADA column)
TARGET_NATUREZA = "ROUBO - OUTROS"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Known SP Capital delegacias (from 2022-04 Wayback snapshot of Pesquisa.aspx)
# value = internal SSP dropdown ID, name = display text
# ---------------------------------------------------------------------------
SP_CAPITAL_DPS_FROM_ASPX = {
    1410: "001 DP - Sé",
    1246: "002 DP - Bom Retiro",
    1143: "003 DP - Campos Elísios",
    1067: "004 DP - Consolação",
    1015: "005 DP - Aclimação",
    983: "006 DP - Cambuci",
    957: "007 DP - Lapa",
    934: "008 DP - Brás",
    915: "009 DP - Carandiru",
    1275: "010 DP - Penha de França",
    1265: "011 DP - Santo Amaro",
    1262: "012 DP - Pari",
    1259: "013 DP - Casa Verde",
    1257: "014 DP - Pinheiros",
    1256: "015 DP - Itaim Bibi",
    1255: "016 DP - Vila Clementino",
    1254: "017 DP - Ipiranga",
    1253: "018 DP - Alto da Moóca",
    1252: "019 DP - Vila Maria",
    1154: "020 DP - Água Fria",
    1153: "021 DP - Vila Matilde",
    1152: "022 DP - São Miguel Paulista",
    1151: "023 DP - Perdizes",
    1150: "024 DP - Ponte Rasa",
    1149: "025 DP - Parelheiros",
    1148: "026 DP - Sacomã",
    1147: "027 DP - Campo Belo",
    1146: "028 DP - Freguesia do Ó",
    1145: "029 DP - Vila Diva",
    1078: "030 DP - Tatuapé",
    1077: "031 DP - Vila Carrão",
    1076: "032 DP - Itaquera",
    1075: "033 DP - Pirituba",
    1074: "034 DP - Vila Sônia",
    1073: "035 DP - Jabaquara",
    1072: "036 DP - Vila Mariana",
    1071: "037 DP - Campo Limpo",
    1070: "038 DP - Vila Amália",
    1069: "039 DP - Vila Gustavo",
    1027: "040 DP - Vila Santa Maria",
    1026: "041 DP - Vila Rica",
    1025: "042 DP - Parque São Lucas",
    1024: "043 DP - Cidade Ademar",
    1023: "044 DP - Guaianazes",
    1022: "045 DP - Vila Brasilândia",
    1021: "046 DP - Perus",
    1020: "047 DP - Capão Redondo",
    1019: "048 DP - Cidade Dutra",
    1018: "049 DP - São Mateus",
    994: "050 DP - Itaim Paulista",
    993: "051 DP - Butantã",
    992: "052 DP - Parque São Jorge",
    991: "053 DP - Parque do Carmo",
    990: "054 DP - Cidade Tiradentes",
    989: "055 DP - Parque São Rafael",
    988: "056 DP - Vila Alpina",
    987: "057 DP - Parque da Moóca",
    986: "058 DP - Vila Formosa",
    985: "059 DP - Jardim Noemia",
    966: "062 DP - Ermelino Matarazzo",
    965: "063 DP - Vila Jacuí",
    964: "064 DP - Cidade A E Carvalho",
    963: "065 DP - Artur Alvim",
    962: "066 DP - Vale do Aricanduva",
    961: "067 DP - Jardim Robru",
    960: "068 DP - Lajeado",
    959: "069 DP - Teotônio Vilela",
    945: "070 DP - Vila Ema",
    943: "072 DP - Vila Penteado",
    942: "073 DP - Jaçanã",
    941: "074 DP - Jaraguá",
    940: "075 DP - Jardim Arpoador",
    938: "077 DP - Santa Cecília",
    937: "078 DP - Jardins",
    926: "080 DP - Vila Joaniza",
    925: "081 DP - Belém",
    923: "083 DP - Parque Bristol",
    921: "085 DP - Jardim Mirna",
    919: "087 DP - Vila Pereira Barreto",
    917: "089 DP - Portal do Morumbi",
    910: "090 DP - Parque Novo Mundo",
    909: "091 DP - Ceagesp",
    908: "092 DP - Parque Santo Antônio",
    907: "093 DP - Jaguaré",
    905: "095 DP - Heliópolis",
    904: "096 DP - Monções",
    903: "097 DP - Americanópolis",
    902: "098 DP - Jardim Míriam",
    901: "099 DP - Campo Grande",
    1270: "100 DP - Jardim Herculano",
    1269: "101 DP - Jardim das Imbuias",
    1268: "102 DP - Socorro",
    1267: "103 DP - Cohab Itaquera",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def fetch_manifest() -> dict:
    """
    Fetches the JSON manifest listing available XLSX file URLs.
    Returns dict mapping year (int) to list of remote file paths.
    """
    log.info("Fetching manifest from %s", MANIFEST_URL)
    req = urllib.request.Request(
        MANIFEST_URL,
        headers={"User-Agent": "Mozilla/5.0 (compatible; SSP-Scraper/1.0)"},
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        raw = resp.read().decode("utf-8-sig")
    data = json.loads(raw)

    year_files: dict[int, list[str]] = defaultdict(list)
    for category in data.get("data", []):
        if "criminal" not in category.get("nome", "").lower():
            continue
        for entry in category.get("lista", []):
            try:
                year = int(str(entry["periodo"]).strip())
            except ValueError:
                continue
            year_files[year].append(entry["arquivo"])
    return year_files


def download_file(remote_path: str, local_path: str) -> None:
    """Downloads a file from the SSP portal, with progress logging."""
    url = f"{BASE_URL}/{remote_path}"
    log.info("Downloading %s → %s", url, local_path)
    os.makedirs(os.path.dirname(local_path), exist_ok=True)
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "Mozilla/5.0 (compatible; SSP-Scraper/1.0)"},
    )
    with urllib.request.urlopen(req, timeout=120) as resp:
        data = resp.read()
    with open(local_path, "wb") as f:
        f.write(data)
    log.info("  → %d MB", len(data) // (1024 * 1024))


def get_local_path(remote_path: str) -> str:
    filename = os.path.basename(remote_path)
    return os.path.join(CACHE_DIR, filename)


# ---------------------------------------------------------------------------
# Core aggregation
# ---------------------------------------------------------------------------

def aggregate_xlsx(path: str) -> tuple[
    dict[tuple[str, int, int], int],   # (dp_circ, year, month) → count
    set[str],                           # all DECAP delegacia names seen
]:
    """
    Reads an SPDadosCriminais XLSX and aggregates ROUBO - OUTROS counts
    by (delegacia_circunscrição, ano_estatistica, mes_estatistica) for DECAP only.

    Returns:
      counts dict  – key: (delegacia_circ_name, year, month)
      decap_dp_set – set of unique delegacia names seen in DECAP rows
    """
    log.info("Processing %s …", os.path.basename(path))
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    counts: dict[tuple[str, int, int], int] = defaultdict(int)
    dp_set: set[str] = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        log.info("  Sheet: %s", sheet_name)
        header = None
        row_count = 0

        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = {str(v).strip(): i for i, v in enumerate(row) if v is not None}
                # Validate required columns
                required = {"NOME_DEPARTAMENTO", "NATUREZA_APURADA",
                            "NOME_DELEGACIA_CIRCUNSCRIÇÃO", "NOME_DELEGACIA",
                            "ANO_ESTATISTICA", "MES_ESTATISTICA"}
                missing = required - set(header.keys())
                if missing:
                    log.warning("    Missing columns: %s – skipping sheet", missing)
                    break
                continue

            dept = str(row[header["NOME_DEPARTAMENTO"]] or "").strip()
            if DECAP not in dept:
                continue

            row_count += 1
            natureza = str(row[header["NATUREZA_APURADA"]] or "").strip()
            circ = str(row[header.get("NOME_DELEGACIA_CIRCUNSCRIÇÃO", -1) if "NOME_DELEGACIA_CIRCUNSCRIÇÃO" in header else -1] or "").strip()
            dp_nome = str(row[header["NOME_DELEGACIA"]] or "").strip()

            # Use circunscrição name (where crime happened) if available
            dp_key = circ if circ and circ != "None" else dp_nome
            dp_key = dp_key.strip()

            if dp_key:
                dp_set.add(dp_key)

            if TARGET_NATUREZA in natureza:
                try:
                    ano = int(row[header["ANO_ESTATISTICA"]])
                    mes_raw = row[header["MES_ESTATISTICA"]]
                    mes = int(mes_raw) if mes_raw is not None else 0
                except (TypeError, ValueError):
                    ano, mes = 0, 0
                counts[(dp_key, ano, mes)] += 1

        log.info("    DECAP rows processed: %d", row_count)

    wb.close()
    return dict(counts), dp_set


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    os.makedirs(CACHE_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # ---- fetch manifest ------------------------------------------------
    try:
        year_files = fetch_manifest()
        log.info("Manifest: years available = %s", sorted(year_files.keys()))
    except Exception as e:
        log.error("Could not fetch manifest: %s", e)
        log.info("Falling back to hard-coded file list …")
        year_files = {
            y: [f"assets/estatistica/transparencia/spDados/SPDadosCriminais_{y}.xlsx"]
            for y in [2022, 2023, 2024, 2025]
        }

    # ---- download files ------------------------------------------------
    local_files: list[tuple[int, str]] = []
    for year in sorted(year_files.keys()):
        if year not in TARGET_YEARS:
            continue
        for remote_path in year_files[year]:
            lpath = get_local_path(remote_path)
            if DOWNLOAD_FRESH or not os.path.exists(lpath):
                try:
                    download_file(remote_path, lpath)
                    time.sleep(1)  # be polite
                except Exception as e:
                    log.error("Download failed for %s: %s", remote_path, e)
                    continue
            local_files.append((year, lpath))

    if not local_files:
        log.error("No files to process. Exiting.")
        return

    # ---- aggregate ------------------------------------------------
    # monthly: (dp, year, month) → count
    monthly: dict[tuple[str, int, int], int] = defaultdict(int)
    all_decap_dps: set[str] = set()

    for year, path in local_files:
        counts, dp_set = aggregate_xlsx(path)
        for k, v in counts.items():
            monthly[k] += v
        all_decap_dps.update(dp_set)

    # ---- annual roll-up ------------------------------------------------
    annual: dict[tuple[str, int], int] = defaultdict(int)
    for (dp, year, month), count in monthly.items():
        annual[(dp, year)] += count

    # ---- write delegacia list ----------------------------------------
    dp_list_path = os.path.join(OUTPUT_DIR, "sp_capital_delegacias.csv")
    with open(dp_list_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["delegacia_xlsx_name", "delegacia_aspx_name", "aspx_id"])

        # Build lookup from aspx names (normalised)
        aspx_lookup = {v.upper(): (k, v) for k, v in SP_CAPITAL_DPS_FROM_ASPX.items()}

        for dp in sorted(all_decap_dps):
            norm = dp.upper()
            aspx_match = aspx_lookup.get(norm, (None, ""))
            w.writerow([dp, aspx_match[1], aspx_match[0] or ""])

    log.info("Wrote %d delegacias → %s", len(all_decap_dps), dp_list_path)

    # ---- write annual CSV -------------------------------------------
    annual_path = os.path.join(OUTPUT_DIR, "roubo_outros_por_delegacia.csv")
    years_present = sorted({y for _, y in annual.keys()})

    with open(annual_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["delegacia_name", "year", "roubo_outros",
                    "note"])
        for (dp, year), count in sorted(annual.items(), key=lambda x: (x[0][0], x[0][1])):
            if year in TARGET_YEARS:
                w.writerow([
                    dp, year, count,
                    "Source: SSP-SP SPDadosCriminais XLSX; "
                    "NATUREZA_APURADA=ROUBO - OUTROS; "
                    "NOME_DEPARTAMENTO=DECAP; "
                    "Counts from NOME_DELEGACIA_CIRCUNSCRIÇÃO"
                ])

    log.info("Wrote annual CSV → %s", annual_path)

    # ---- write monthly CSV -------------------------------------------
    monthly_path = os.path.join(OUTPUT_DIR, "roubo_outros_mensal.csv")
    with open(monthly_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["delegacia_name", "year", "month", "roubo_outros"])
        for (dp, year, month), count in sorted(monthly.items()):
            if year in TARGET_YEARS:
                w.writerow([dp, year, month, count])

    log.info("Wrote monthly CSV → %s", monthly_path)

    # ---- print sample -----------------------------------------------
    print("\n=== SAMPLE: Top 20 delegacias by roubo_outros (2022) ===")
    yr2022 = [(dp, c) for (dp, yr), c in annual.items() if yr == 2022]
    yr2022.sort(key=lambda x: -x[1])
    print(f"{'Delegacia':<45} {'Count':>6}")
    print("-" * 55)
    for dp, c in yr2022[:20]:
        print(f"{dp:<45} {c:>6}")

    print("\n=== Annual totals for SP capital ===")
    for y in sorted(years_present):
        if y in TARGET_YEARS:
            total = sum(c for (_, yr), c in annual.items() if yr == y)
            n_dp = len({dp for (dp, yr) in annual if yr == y})
            print(f"  {y}: {total:6d}  roubo_outros across {n_dp} delegacias")


if __name__ == "__main__":
    main()
